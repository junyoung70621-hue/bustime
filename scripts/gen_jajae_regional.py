# -*- coding: utf-8 -*-
# ─────────────────────────────────────────────────────────────
# 대전·세종 자재 지급확인서 엑셀 → lib/daepyecha-regional/items.generated.ts
#   시트 6개: B400/B500/B650 × (대폐차 / 증차). 각 시트 8행~"기타" 직전까지 품목.
#   name=B열, bigo=D열("* " 제거), hasNewReused = bigo에 "(신규" 포함. 기본수량=대당 1.
#   재생성: python scripts/gen_jajae_regional.py [xlsx경로]
# ─────────────────────────────────────────────────────────────
import sys, os, json
import openpyxl

DEFAULT_XLSX = os.path.expanduser(r"~/Downloads/자재지급확인서_양식_대전_세종 (1).xlsx")
XLSX = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

MODELS = ["B400", "B500", "B650"]

# 엑셀엔 있으나 실제 양식에서 제외하는 품목(이름 기준). 사용자 지시 반영.
#   - "3S 앰프": B400 증차 시트에 있으나 제외(증차도 대폐차와 동일 15품목).
EXCLUDE_NAMES = {"3S 앰프"}


def norm(v):
    return "" if v is None else str(v).replace("\r\n", "\n").strip()


def extract(ws):
    a1 = norm(ws.cell(1, 1).value)
    model = next((m for m in MODELS if a1.startswith(m) or a1.lstrip().startswith(m)), None)
    a4 = norm(ws.cell(4, 1).value)  # "2. 용   도 : 대폐차"
    purpose = "증차" if "증차" in a4 else "대폐차"
    # 헤더(NO/품목) 행 찾기
    hr = None
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value) == "NO":
            hr = r
            break
    items = []
    for r in range(hr + 1, ws.max_row + 1):
        name = norm(ws.cell(r, 2).value)  # B열
        if name == "" or name == "기타":
            if name == "기타":
                break
            continue
        if name in EXCLUDE_NAMES:
            continue
        bigo = norm(ws.cell(r, 4).value)  # D열
        if bigo.startswith("*"):
            bigo = bigo[1:].strip()
        has_nr = "(신규" in bigo
        items.append({"name": name, "bigo": bigo, "hasNewReused": has_nr})
    return model, purpose, items


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data = {m: {} for m in MODELS}
    for ws in wb.worksheets:
        model, purpose, items = extract(ws)
        if not model:
            continue
        data[model][purpose] = items

    out = []
    out.append("// ─────────────────────────────────────────────────────────────")
    out.append("// [자동 생성] 대전·세종 자재 지급확인서 모델×용도별 품목")
    out.append("//   생성: python scripts/gen_jajae_regional.py")
    out.append("//   원본: 자재지급확인서_양식_대전_세종 (1).xlsx")
    out.append("// ─────────────────────────────────────────────────────────────")
    out.append('import type { ItemTemplate } from "@/lib/daepyecha/types";')
    out.append("")
    out.append('export type RegJajaeModel = "B400" | "B500" | "B650";')
    out.append('export type RegJajaePurpose = "대폐차" | "증차";')
    out.append('export const REG_JAJAE_MODELS: RegJajaeModel[] = ["B400", "B500", "B650"];')
    out.append("")
    out.append("export const REG_JAJAE_ITEMS: Record<RegJajaeModel, Record<RegJajaePurpose, ItemTemplate[]>> = {")
    for m in MODELS:
        out.append(f"  {m}: {{")
        for purpose in ["대폐차", "증차"]:
            items = data[m].get(purpose, [])
            out.append(f'    "{purpose}": [')
            for it in items:
                out.append(
                    "      { "
                    + f'name: {json.dumps(it["name"], ensure_ascii=False)}, '
                    + f'bigo: {json.dumps(it["bigo"], ensure_ascii=False)}, '
                    + f'hasNewReused: {"true" if it["hasNewReused"] else "false"} }},'
                )
            out.append("    ],")
        out.append("  },")
    out.append("};")
    out.append("")

    ts_path = os.path.join(os.path.dirname(__file__), "..", "lib", "daepyecha-regional", "items.generated.ts")
    os.makedirs(os.path.dirname(ts_path), exist_ok=True)
    with open(ts_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print("items.generated.ts 생성 완료")
    for m in MODELS:
        for p in ["대폐차", "증차"]:
            print(f"  {m} {p}: {len(data[m].get(p, []))} 품목")


if __name__ == "__main__":
    main()
