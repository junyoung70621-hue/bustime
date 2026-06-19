# -*- coding: utf-8 -*-
# ─────────────────────────────────────────────────────────────
# 대전·세종 설치완료 체크리스트 엑셀 → lib/checklist-regional/models.generated.ts 생성
#   원본: "00.대폐차 설치완료 체크리스트(1123)_대전_세종_20201123.xlsx"
#   시트 4개(B400/B500/B600/B650). 8열표(케이스/점검대상/점검항목/점검방법/점검POINT/O,X/비고).
#   - 케이스(A)·케이스명(B)은 B열(케이스명) 블록 기준으로 정규화(엑셀 A열 병합이 불완전).
#   - C~H는 원본 병합 그대로 m:{컬럼→[rowspan,colspan]} 로 반영.
#   - kind: vehicleType(차량특성 종류) / partition(격벽) / etc(특이사항) / check(나머지).
#   재생성: python scripts/gen_checklist_regional.py [xlsx경로]
# ─────────────────────────────────────────────────────────────
import sys, os, json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

DEFAULT_XLSX = os.path.expanduser(
    r"~/Downloads/00.대폐차 설치완료 체크리스트(1123)_대전_세종_20201123.xlsx"
)
XLSX = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

# 논리 컬럼 = (엑셀 열문자, 키)
COLS = [
    ("A", "caseNo"), ("B", "caseLabel"), ("C", "target"), ("D", "item"),
    ("E", "method"), ("F", "point"), ("G", "ox"), ("H", "bigo"),
]

MODELS = ["B400", "B500", "B600", "B650"]


def norm(v):
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").strip()


def build_merge_index(ws):
    """covered[(r,c)] = (top_r, top_c);  span[(r,c)] = (rows, cols)  (top-left만)"""
    covered = {}
    span = {}
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        span[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    covered[(r, c)] = (min_r, min_c)
    return covered, span


def detect_header_row(ws):
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value) == "케이스":
            return r
    raise RuntimeError("헤더행(케이스) 탐색 실패")


def cell_val(ws, covered, r, c):
    """병합에 가려진 셀은 top-left 값을 반환"""
    if (r, c) in covered:
        tr, tc = covered[(r, c)]
        return norm(ws.cell(tr, tc).value)
    return norm(ws.cell(r, c).value)


def kind_for(target, item, method, point, case_label):
    p = point + " " + method + " " + item
    if "특이사항" in case_label:
        return "etc"
    if "격벽설치 유무" in p:
        return "partition"
    if ("현대" in p and "저상" in p) or "차량 특성 종류" in p:
        return "vehicleType"
    return "check"


def extract_sheet(ws):
    covered, span = build_merge_index(ws)
    hr = detect_header_row(ws)
    bcol = column_index_from_string("B")

    body_rows = list(range(hr + 1, ws.max_row + 1))

    # B열(케이스명) top-left 위치로 케이스 블록 분할
    blocks = []  # (start_row, end_row, label, label_colspan)
    cur = None
    for r in body_rows:
        b_is_top = (r, bcol) not in covered and norm(ws.cell(r, bcol).value) != ""
        if b_is_top:
            if cur:
                cur[1] = r - 1
                blocks.append(cur)
            label = norm(ws.cell(r, bcol).value)
            # B 병합 colspan(예: 차량특성은 B:D 3칸)
            colspan = span.get((r, bcol), (1, 1))[1]
            cur = [r, None, label, colspan]
    if cur:
        cur[1] = body_rows[-1]
        blocks.append(cur)

    # 행→(블록인덱스, 블록) 매핑
    row_block = {}
    for i, blk in enumerate(blocks):
        for r in range(blk[0], blk[1] + 1):
            row_block[r] = (i, blk)

    rows_out = []
    for r in body_rows:
        bi, blk = row_block[r]
        case_no = bi + 1
        is_block_top = (r == blk[0])
        label_colspan = blk[3]

        vals = {}
        m = {}
        for letter, key in COLS:
            c = column_index_from_string(letter)
            if key == "caseNo":
                # 정규화: 블록 첫 행에만 caseNo, rowspan=블록크기
                if is_block_top:
                    vals[key] = str(case_no)
                    m[key] = [blk[1] - blk[0] + 1, 1]
                continue
            if key == "caseLabel":
                if is_block_top:
                    vals[key] = blk[2]
                    m[key] = [blk[1] - blk[0] + 1, label_colspan]
                continue
            # caseLabel colspan 에 가려지는 target/item 등은 건너뜀
            if is_block_top and label_colspan > 1 and c in range(bcol + 1, bcol + label_colspan):
                continue
            if (r, c) in covered:
                # caseLabel 정규화로 가려진 행(블록 첫행 외) 처리: B 블록이 colspan>1이면 해당 열 skip
                tr, tc = covered[(r, c)]
                if tc == bcol and label_colspan > 1:
                    continue  # caseLabel 병합에 흡수
                continue  # 다른 병합에 가려짐 → 렌더 안 함
            # top-left(또는 단일) → 렌더
            sp = span.get((r, c), (1, 1))
            vals[key] = cell_val(ws, covered, r, c)
            m[key] = [sp[0], sp[1]]

        kind = kind_for(
            vals.get("target", ""), vals.get("item", ""),
            vals.get("method", ""), vals.get("point", ""), blk[2],
        )
        rows_out.append({
            "caseNo": case_no,
            "caseLabel": vals.get("caseLabel", "") if is_block_top else "",
            "target": vals.get("target", ""),
            "item": vals.get("item", ""),
            "method": vals.get("method", ""),
            "point": vals.get("point", ""),
            "ox": vals.get("ox", ""),
            "kind": kind,
            "bigo": vals.get("bigo", ""),
            "bigoKind": "static",
            "m": m,
        })
    title = norm(ws.cell(1, 1).value)
    return {"title": title, "rows": rows_out}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    data = {}
    for ws in wb.worksheets:
        # 시트명 깨짐 → 본문 A1("B400 단말기 체크 리스트")에서 모델 식별
        a1 = norm(ws.cell(1, 1).value)
        model = next((m for m in MODELS if a1.startswith(m)), None)
        if not model:
            continue
        data[model] = extract_sheet(ws)

    # 리뷰용 JSON
    with open(os.path.join(os.path.dirname(__file__), "..", "_ck_regional_review.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    # TS 출력
    out = []
    out.append("// ─────────────────────────────────────────────────────────────")
    out.append("// [자동 생성] 대전·세종 설치완료 체크리스트 모델별 점검표")
    out.append("//   생성: python scripts/gen_checklist_regional.py")
    out.append("//   원본: 00.대폐차 설치완료 체크리스트(1123)_대전_세종_20201123.xlsx")
    out.append("//   수정 시 이 파일을 직접 고치지 말고 스크립트/엑셀을 고쳐 재생성하세요.")
    out.append("// ─────────────────────────────────────────────────────────────")
    out.append('import type { CkRowDef } from "@/lib/checklist/types";')
    out.append("")
    out.append("export type RegModel = " + " | ".join(f'"{m}"' for m in MODELS) + ";")
    out.append("export const REG_MODELS: RegModel[] = [" + ", ".join(f'"{m}"' for m in MODELS) + "];")
    out.append("")
    out.append("export type RegModelDef = { title: string; rows: CkRowDef[] };")
    out.append("export const REG_MODELS_DATA: Record<RegModel, RegModelDef> = {")
    for model in MODELS:
        d = data[model]
        out.append(f'  {model}: {{')
        out.append(f'    title: {json.dumps(d["title"], ensure_ascii=False)},')
        out.append("    rows: [")
        for r in d["rows"]:
            m_str = "{ " + ", ".join(
                f'{k}: [{v[0]}, {v[1]}]' for k, v in r["m"].items()
            ) + " }"
            out.append(
                "      { "
                + f'caseNo: {r["caseNo"]}, '
                + f'caseLabel: {json.dumps(r["caseLabel"], ensure_ascii=False)}, '
                + f'target: {json.dumps(r["target"], ensure_ascii=False)}, '
                + f'item: {json.dumps(r["item"], ensure_ascii=False)}, '
                + f'method: {json.dumps(r["method"], ensure_ascii=False)}, '
                + f'point: {json.dumps(r["point"], ensure_ascii=False)}, '
                + f'kind: {json.dumps(r["kind"])}, '
                + f'bigo: {json.dumps(r["bigo"], ensure_ascii=False)}, '
                + f'bigoKind: "static", '
                + f'm: {m_str} }},'
            )
        out.append("    ],")
        out.append("  },")
    out.append("};")
    out.append("")

    ts_path = os.path.join(os.path.dirname(__file__), "..", "lib", "checklist-regional", "models.generated.ts")
    os.makedirs(os.path.dirname(ts_path), exist_ok=True)
    with open(ts_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print("models.generated.ts 생성 완료:", len(data), "모델")
    for m in MODELS:
        print(" ", m, "rows:", len(data[m]["rows"]))


if __name__ == "__main__":
    main()
